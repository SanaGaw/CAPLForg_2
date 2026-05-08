"""Excel test plan parser for CAPL Pipeline V2.2.

Parses Capgemini-format Excel test plans to extract test cases,
test steps, and their associated signals and parameters.
"""

from pathlib import Path
from typing import Dict, List, Optional
from dataclasses import dataclass, field
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass
class TestStep:
    """Represents a single test step within a test case."""
    step_id: str  # e.g., "TC_001", "1.2.3"
    step_number: int
    description: str
    action: Optional[str] = None
    expected_result: Optional[str] = None
    signal_refs: List[str] = field(default_factory=list)
    parameters: Dict[str, str] = field(default_factory=dict)
    excel_row: int = 0
    notes: Optional[str] = None


@dataclass
class TestCase:
    """Represents a test case with its steps."""
    test_id: str
    title: str
    description: Optional[str] = None
    category: Optional[str] = None
    priority: Optional[str] = None
    signals: List[str] = field(default_factory=list)
    steps: List[TestStep] = field(default_factory=list)
    excel_sheet: Optional[str] = None
    source_file: str = ""


class ExcelParser:
    """
    Parse Capgemini-format Excel test plans.

    Expected format:
    - Column headers in first row
    - Test case data starts from second row
    - Steps may be in sub-rows or separate columns

    Supports xlsx and xls formats via openpyxl.
    """

    # Default column mappings (adjustable)
    DEFAULT_COLUMNS = {
        'test_id': ['Test ID', 'TestCase ID', 'TC_ID', 'ID'],
        'title': ['Title', 'Test Title', 'Description', 'Test Description'],
        'category': ['Category', 'Type', 'Test Type'],
        'priority': ['Priority', 'Prio'],
        'step_number': ['Step', 'Step No', 'Step #'],
        'action': ['Action', 'Test Action', 'Step Action'],
        'expected': ['Expected Result', 'Expected', 'Verification'],
        'signal': ['Signal', 'Signals', 'Signal Name', 'Signal Reference'],
    }

    def __init__(self, excel_path: Path) -> None:
        self.excel_path = excel_path
        self._workbook: Optional[Workbook] = None
        self.test_cases: Dict[str, TestCase] = {}

    def _load_workbook(self) -> Workbook:
        """Load and cache the Excel workbook."""
        if self._workbook is None:
            self._workbook = load_workbook(
                self.excel_path,
                data_only=True,  # Read values, not formulas
                read_only=False
            )
        return self._workbook

    def parse(self, sheet_name: Optional[str] = None) -> Dict[str, TestCase]:
        """
        Parse Excel file and return test cases.

        Args:
            sheet_name: Specific sheet to parse. If None, parses all sheets.

        Returns:
            Dict mapping test IDs to TestCase objects
        """
        wb = self._load_workbook()
        self.test_cases = {}

        sheets_to_parse = [wb[sheet_name]] if sheet_name else wb.worksheets

        for worksheet in sheets_to_parse:
            self._parse_sheet(worksheet)

        logger.info(
            f"Parsed {self.excel_path.name}: "
            f"{len(self.test_cases)} test cases"
        )
        return self.test_cases

    def _parse_sheet(self, sheet: Worksheet) -> None:
        """Parse a single worksheet."""
        # Read all rows
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        # Determine column mapping from header row
        header_row = rows[0]
        col_map = self._detect_columns(header_row)

        if not col_map:
            logger.warning(f"No recognized columns in sheet {sheet.title}")
            return

        # Parse data rows
        current_test_case: Optional[TestCase] = None
        step_number = 0

        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue  # Skip empty rows

            # Check if this is a new test case
            test_id_col = col_map.get('test_id')
            if test_id_col is not None and row[test_id_col]:
                test_id = str(row[test_id_col]).strip()
                if test_id:
                    title_col = col_map.get('title')
                    title = str(row[title_col]) if title_col is not None and row[title_col] else test_id

                    current_test_case = TestCase(
                        test_id=test_id,
                        title=title,
                        description=str(row[title_col]) if title_col is not None and row[title_col] else None,
                        excel_sheet=sheet.title,
                        source_file=str(self.excel_path)
                    )
                    self.test_cases[test_id] = current_test_case
                    step_number = 0

            # If we have a current test case, parse the step
            if current_test_case:
                step_number += 1
                step = self._parse_step(row, col_map, row_idx, step_number)
                if step:
                    current_test_case.steps.append(step)

                    # Collect signals from step
                    for sig in step.signal_refs:
                        if sig not in current_test_case.signals:
                            current_test_case.signals.append(sig)

    def _detect_columns(self, header_row: tuple) -> Dict[str, int]:
        """Detect column indices from header row."""
        col_map: Dict[str, int] = {}

        for col_idx, header in enumerate(header_row):
            if header is None:
                continue

            header_str = str(header).strip()

            for col_name, possible_headers in self.DEFAULT_COLUMNS.items():
                if header_str in possible_headers:
                    col_map[col_name] = col_idx
                    break

        return col_map

    def _parse_step(
        self,
        row: tuple,
        col_map: Dict[str, int],
        excel_row: int,
        step_number: int
    ) -> Optional[TestStep]:
        """Parse a single step from a row."""
        step_id = ""
        test_id_col = col_map.get('test_id')
        if test_id_col is not None and row[test_id_col]:
            step_id = f"{row[test_id_col]}.{step_number}"

        # Get description/action
        action_col = col_map.get('action')
        action = str(row[action_col]) if action_col is not None and row[action_col] else None

        expected_col = col_map.get('expected')
        expected = str(row[expected_col]) if expected_col is not None and row[expected_col] else None

        # Get signals
        signal_col = col_map.get('signal')
        signal_refs: List[str] = []
        if signal_col is not None and row[signal_col]:
            signals_str = str(row[signal_col])
            # Split by common delimiters
            for sig in signals_str.replace(';', ',').split(','):
                sig = sig.strip()
                if sig:
                    signal_refs.append(sig)

        # Build description from action or expected
        description = action or expected or f"Step {step_number}"
        if expected and action:
            description = f"{action} -> {expected}"
        elif expected:
            description = expected

        return TestStep(
            step_id=step_id,
            step_number=step_number,
            description=description,
            action=action,
            expected_result=expected,
            signal_refs=signal_refs,
            excel_row=excel_row
        )

    def get_test_case(self, test_id: str) -> Optional[TestCase]:
        """Get a test case by ID."""
        if not self.test_cases:
            self.parse()
        return self.test_cases.get(test_id)

    def get_total_steps(self) -> int:
        """Get total number of steps across all test cases."""
        if not self.test_cases:
            self.parse()
        return sum(len(tc.steps) for tc in self.test_cases.values())

    def get_all_signals(self) -> List[str]:
        """Get all unique signal references."""
        if not self.test_cases:
            self.parse()
        signals: List[str] = []
        for tc in self.test_cases.values():
            for sig in tc.signals:
                if sig not in signals:
                    signals.append(sig)
        return signals
